
import streamlit as st
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

st.set_page_config(page_title="Formularz Katalogu Produktów", layout="wide")
st.title("📦 Formularz Katalogu Produktów")

UPLOAD_DIR = "uploads"
EXCEL_FILE = "katalog_web.xlsx"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 🧹 Czyść formularz
if st.button("🧹 Czyść formularz"):
    st.session_state.clear()
    st.success("Formularz został wyczyszczony. Odśwież stronę (F5), aby wczytać czysty formularz.")
    st.stop()

# 📤 Formularz wprowadzania danych
with st.form("formularz"):
    col1, col2 = st.columns(2)
    with col1:
        nazwa = st.text_input("Nazwa produktu")
        kod = st.text_input("Kod produktu (SKU)")
        wersja = st.text_input("Wersja / Rewizja")
        typ = st.selectbox("Typ", ["Butelka", "Słoik", "Nakrętka", "Nasadka", "Dysktop", "Fliptop", "Inne"])
        gwint = st.text_input("Gwint")
        material = st.selectbox("Materiał", ["PET", "PP", "HDPE", "PETG"])
        technologia = st.selectbox("Technologia produkcji", ["IM", "ISBM", "EBM"])
        kolor = st.text_input("Kolorystyka")
        wymiary = st.text_input("Wymiary (Ø x H)")
        masa = st.text_input("Masa (g)")
    with col2:
        pakowanie = st.selectbox("Pakowanie", ["Luz", "Układane"])
        typ_opakowania = st.selectbox("Typ opakowania", [
            "K1", "K2", "K3", "K4", "K5", "K6", "K7", "K8", "K9", "K10", "K11", "K12", "K13",
            "T1 800x600", "T2 1000x600"])
        opakowanie = st.text_input("Opakowanie (szt/karton, szt/paleta)")
        certyfikaty = st.text_input("Certyfikaty")
        montaz = st.selectbox("Montaż", ["Tak", "Nie"])
        wsp_element = st.text_input("Element współpracujący") if montaz == "Tak" else ""
        rysunek = st.file_uploader("Plik rysunek")
        spec = st.file_uploader("Plik specyfikacja")
        schemat = st.file_uploader("Schemat pakowania")
        uwagi = st.text_area("Uwagi")

    submitted = st.form_submit_button("💾 Zapisz")

# 🔽 Zapis danych
if submitted:
    def save_file(file):
        if file:
            filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.name}"
            path = os.path.join(UPLOAD_DIR, filename)
            with open(path, "wb") as f:
                f.write(file.getbuffer())
            return path
        return ""

    rysunek_path = save_file(rysunek)
    spec_path = save_file(spec)
    schemat_path = save_file(schemat)

    row_data = [
        nazwa, kod, wersja, typ, gwint, material, technologia, kolor, wymiary, masa,
        pakowanie, typ_opakowania, opakowanie, certyfikaty, montaz, wsp_element,
        rysunek_path, spec_path, schemat_path,
        datetime.today().strftime("%Y-%m-%d"), uwagi
    ]

    headers = [
        "Nazwa produktu", "Kod produktu (SKU)", "Wersja / Rewizja", "Typ", "Gwint",
        "Materiał", "Technologia produkcji", "Kolorystyka", "Wymiary (Ø x H)", "Masa (g)",
        "Pakowanie", "Typ opakowania", "Opakowanie (szt/karton, szt/paleta)", "Certyfikaty",
        "Montaż", "Element współpracujący", "Plik rysunek", "Plik specyfikacja", "Schemat pakowania",
        "Data modyfikacji", "Uwagi"
    ]

    if not os.path.exists(EXCEL_FILE):
        pd.DataFrame([row_data], columns=headers).to_excel(EXCEL_FILE, index=False)
    else:
        book = load_workbook(EXCEL_FILE)
        sheet = book.active
        row_idx = sheet.max_row + 1
        for col, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col)
            if col in [17, 18, 19] and isinstance(value, str) and os.path.exists(value):
                name = os.path.basename(value)
                cell.value = name
                cell.hyperlink = value
                cell.font = Font(underline='single', color="0000FF")
            else:
                cell.value = value
        book.save(EXCEL_FILE)

    st.success("✅ Dane i hiperłącza do plików zostały zapisane!")

# 👀 Podgląd istniejących wpisów
st.markdown("---")
st.subheader("📄 Podgląd zapisanych produktów")

if os.path.exists(EXCEL_FILE):
    df_show = pd.read_excel(EXCEL_FILE)
    st.dataframe(df_show, use_container_width=True)
else:
    st.info("Brak jeszcze zapisanych produktów w katalogu.")
